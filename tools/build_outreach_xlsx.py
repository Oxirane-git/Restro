"""
Builds a 2-sheet Excel file:
  Sheet 1: Cafes     (unique emails, no overlap with restaurants)
  Sheet 2: Restaurants (unique emails, no overlap with cafes)

Tiebreak rule: if the same email appears in BOTH categories,
it goes to RESTAURANTS (higher-value niche typically).
"""

import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LEADS_DIR = r"c:\Users\sahil\Desktop\Restro\Leads"
OUTPUT_DIR = r"c:\Users\sahil\Desktop\Restro\Leads"

CAFE_KEYWORDS    = ("cafe", "coffee", "espresso")
RESTO_KEYWORDS   = ("restaurant", "bistro", "diner")

HEADERS = [
    "Business Name", "Email", "Phone", "Website",
    "Owner Name", "Category", "City", "Address", "Source"
]
FIELD_MAP = [
    "business_name", "email", "phone", "website",
    "owner_name", "category", "city", "address", "source"
]

# ── Load all leads ────────────────────────────────────────────────────────────

def load_leads(keyword_tuple):
    leads = {}
    for f in sorted(os.listdir(LEADS_DIR)):
        if not f.endswith(".csv"):
            continue
        if not any(kw in f.lower() for kw in keyword_tuple):
            continue
        path = os.path.join(LEADS_DIR, f)
        try:
            with open(path, "r", encoding="utf-8") as fp:
                for row in csv.DictReader(fp):
                    email = row.get("email", "").strip().lower()
                    if email and email not in leads:
                        leads[email] = {
                            "business_name": row.get("business_name", "").strip(),
                            "email":         email,
                            "phone":         row.get("phone",      "").strip(),
                            "website":       row.get("website",    "").strip(),
                            "owner_name":    row.get("owner_name", "").strip(),
                            "category":      row.get("category",   "").strip(),
                            "city":          row.get("city",       "").strip(),
                            "address":       row.get("address",    "").strip(),
                            "source":        row.get("source",     "google_maps").strip(),
                        }
        except Exception as e:
            print(f"  [WARN] Could not read {f}: {e}")
    return leads

# ── Excel styling ─────────────────────────────────────────────────────────────

def make_header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(ws, rows, theme_hex, accent_hex, sheet_title):
    """Write a leads sheet with formatting."""

    # Column widths
    col_widths = [30, 36, 18, 38, 22, 20, 22, 45, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = make_header_fill(theme_hex)
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = thin_border()

    # ── Data rows ─────────────────────────────────────────────────────────────
    alt_fill  = PatternFill("solid", fgColor="F7F9FC")
    none_fill = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(size=10, name="Calibri")
    link_font = Font(size=10, name="Calibri", color="1155CC", underline="single")

    for row_idx, record in enumerate(rows, 2):
        bg = alt_fill if row_idx % 2 == 0 else none_fill
        ws.row_dimensions[row_idx].height = 18

        for col_idx, field in enumerate(FIELD_MAP, 1):
            value = record.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center")

            # Hyperlink-style for email and website
            if field in ("email", "website") and value:
                cell.font = link_font
            else:
                cell.font = data_font
            cell.fill = bg

    # ── Freeze top row & auto-filter ─────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # ── Summary row at the top (above header) ─────────────────────────────────
    # We'll just set the tab color instead
    ws.sheet_properties.tabColor = accent_hex

    print(f"  [{sheet_title}] {len(rows)} rows written.")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n========================================")
    print("  Building Outreach Excel (2 sheets)")
    print("========================================\n")

    # 1. Load both categories
    print("[1/4] Loading cafe leads...")
    cafe_leads  = load_leads(CAFE_KEYWORDS)
    print(f"      Raw unique cafe emails: {len(cafe_leads)}")

    print("[2/4] Loading restaurant leads...")
    resto_leads = load_leads(RESTO_KEYWORDS)
    print(f"      Raw unique restaurant emails: {len(resto_leads)}")

    # 2. Remove cross-category overlap
    #    Tiebreak: overlap goes to RESTAURANTS (keep in resto, remove from cafe)
    overlap = set(cafe_leads.keys()) & set(resto_leads.keys())
    print(f"\n[3/4] Removing {len(overlap)} cross-category duplicates (kept in Restaurants)...")
    for email in overlap:
        del cafe_leads[email]

    print(f"      Final cafe leads      : {len(cafe_leads)}")
    print(f"      Final restaurant leads: {len(resto_leads)}")
    assert not (set(cafe_leads) & set(resto_leads)), "BUG: overlap still exists!"

    # 3. Sort each sheet: email+phone first, then email-only
    def sort_key(r):
        score = 0
        if r.get("phone"):    score += 4
        if r.get("owner_name"): score += 2
        if r.get("website"):  score += 1
        return -score

    cafe_rows  = sorted(cafe_leads.values(),  key=sort_key)
    resto_rows = sorted(resto_leads.values(), key=sort_key)

    # 4. Write Excel
    print("\n[4/4] Writing Excel file...")
    wb = Workbook()

    # Sheet 1: Cafes  (teal theme)
    ws_cafe = wb.active
    ws_cafe.title = "Cafes"
    write_sheet(ws_cafe, cafe_rows,  theme_hex="1A7A6E", accent_hex="1A7A6E", sheet_title="Cafes")

    # Sheet 2: Restaurants (deep blue theme)
    ws_resto = wb.create_sheet(title="Restaurants")
    write_sheet(ws_resto, resto_rows, theme_hex="1E3A5F", accent_hex="1E3A5F", sheet_title="Restaurants")

    # 5. Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUTPUT_DIR, f"outreach_leads_{timestamp}.xlsx")
    wb.save(out_path)

    # 6. Summary
    c_phone = sum(1 for r in cafe_rows  if r.get("phone"))
    r_phone = sum(1 for r in resto_rows if r.get("phone"))

    print(f"\n========================================")
    print(f"  SAVED: {out_path}")
    print(f"========================================")
    print(f"  Sheet 1 - Cafes:")
    print(f"    Total rows   : {len(cafe_rows)}")
    print(f"    Email + Phone: {c_phone}")
    print(f"  Sheet 2 - Restaurants:")
    print(f"    Total rows   : {len(resto_rows)}")
    print(f"    Email + Phone: {r_phone}")
    print(f"  Grand total    : {len(cafe_rows) + len(resto_rows)} (0 duplicates)")
    print(f"========================================\n")


if __name__ == "__main__":
    main()
